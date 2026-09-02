import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
NHIS_XLSX = ROOT / "source-data" / "nhis-longtermcare-20260610.xlsx"
HIRA_CSV = ROOT / "source-data" / "hira-facilities-20251231.csv"
EVAL_CSV = ROOT / "source-data" / "nhis_longtermcare_evaluations_20260625.csv"
OUT_DIR = ROOT / "nationwide-care-data"
TARGET_CHARS = 155_000

PROVINCES = [
    "서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시",
    "대전광역시", "울산광역시", "세종특별자치시", "경기도", "강원특별자치도",
    "충청북도", "충청남도", "전북특별자치도", "전라남도", "경상북도",
    "경상남도", "제주특별자치도",
]

CATEGORY_CODES = {
    "facility": {"A01", "A02", "A03", "A04", "A05"},
    "daycare": {"B03", "C03"},
    "home-care": {"B01", "C01"},
    "home-nursing": {"B05", "C05"},
    "home-bath": {"B02", "C02"},
    "short-stay": {"B04", "C04"},
}

CATEGORIES = {
    "facility": {"label": "요양시설", "sourceDate": "2026-06-10", "source": "nhis"},
    "daycare": {"label": "주야간보호", "sourceDate": "2026-06-10", "source": "nhis"},
    "home-care": {"label": "방문요양", "sourceDate": "2026-06-10", "source": "nhis"},
    "home-nursing": {"label": "방문간호", "sourceDate": "2026-06-10", "source": "nhis"},
    "home-bath": {"label": "방문목욕", "sourceDate": "2026-06-10", "source": "nhis"},
    "short-stay": {"label": "단기보호", "sourceDate": "2026-06-10", "source": "nhis"},
    "dementia": {"label": "치매전담", "sourceDate": "2026-06-10", "source": "nhis"},
    "nursing-hospital": {"label": "요양병원", "sourceDate": "2025-12-31", "source": "hira"},
}

EVAL_KEYWORDS = {
    "facility": ("입소시설",),
    "daycare": ("주야간보호",),
    "home-care": ("방문요양",),
    "home-nursing": ("방문간호",),
    "home-bath": ("방문목욕",),
    "short-stay": ("단기보호",),
    "dementia": ("입소시설", "주야간보호"),
}


def clean(value):
    return " ".join(str(value or "").split())


def number(value, default=0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def iso_date(value):
    text = re.sub(r"\D", "", str(value or ""))
    return f"{text[:4]}-{text[4:6]}-{text[6:8]}" if len(text) >= 8 else ""


def normalize_id(value):
    return re.sub(r"\D", "", str(value or ""))


def split_region(region, address=""):
    text = clean(region or address)
    parts = text.split()
    province = parts[0] if parts else ""
    if province == "강원도":
        province = "강원특별자치도"
    if province == "전라북도":
        province = "전북특별자치도"
    if province == "세종특별자치시":
        return province, "세종시"
    if len(parts) < 2:
        return province, ""
    city = parts[1]
    if province not in {"서울특별시", "부산광역시", "대구광역시", "인천광역시", "광주광역시", "대전광역시", "울산광역시"}:
        if city.endswith("시") and len(parts) > 2 and parts[2].endswith("구"):
            city += " " + parts[2]
    return province, city


def newest_evaluations():
    rows = defaultdict(dict)
    with EVAL_CSV.open("r", encoding="cp949", newline="") as handle:
        for row in csv.DictReader(handle):
            inst_id = normalize_id(row.get("장기요양기관기호"))
            service = clean(row.get("급여종류"))
            date = clean(row.get("평가일자"))
            date_key = date or "0000-00-00"
            for category, keywords in EVAL_KEYWORDS.items():
                if not any(keyword in service for keyword in keywords):
                    continue
                previous = rows[category].get(inst_id)
                if previous and previous[0] >= date_key:
                    continue
                year_match = re.search(r"(20\d{2})", row.get("평가구분") or "")
                rows[category][inst_id] = (date_key, {
                    "g": clean(row.get("평가등급")),
                    "es": number(row.get("평가총점"), None),
                    "ey": int(year_match.group(1)) if year_match else None,
                    "ed": date,
                    "et": service.split(".", 1)[-1] if "." in service else service,
                })
    return {category: {key: value for key, (_, value) in items.items()} for category, items in rows.items()}


def merge_numeric(target, row):
    target["s"] += number(row[5])
    target["rn"] += number(row[8])
    target["na"] += number(row[9])
    target["pt"] += number(row[11])
    target["ot"] += number(row[12])
    target["cw"] += number(row[13])


def base_record(general, category, evals):
    code, name, _zip, _pc, _cc, _dc, region, designated, _installed, address = general
    province, city = split_region(region, address)
    record = {
        "i": str(code), "n": clean(name), "p": province, "c": city,
        "a": clean(address), "d": iso_date(designated), "t": "", "tn": "",
        "z": 0, "s": 0, "rn": 0, "na": 0, "pt": 0, "ot": 0, "cw": 0,
    }
    record.update(evals.get(category, {}).get(str(code), {}))
    return record


def build_ltc_records():
    wb = openpyxl.load_workbook(NHIS_XLSX, read_only=True, data_only=True)
    general = {str(row[0]): row for row in wb["일반현황"].iter_rows(min_row=2, values_only=True) if row[0]}
    capacities = defaultdict(list)
    for row in wb["입소인원"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            capacities[str(row[0])].append(row)
    staffing = defaultdict(list)
    for row in wb["인력현황"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            staffing[str(row[0])].append(row)

    evals = newest_evaluations()
    output = {category: {} for category in EVAL_KEYWORDS}
    dementia = lambda code: code.startswith(("G", "H", "I", "M")) or code == "S41"

    for category, codes in CATEGORY_CODES.items():
        # 입소인원 시트의 서비스 유형을 운영기관 판정 기준으로 사용한다.
        # 인력 시트는 인력 수치 보강에만 쓰며, 인력 신고만 남은 기관을 중복 포함하지 않는다.
        ids = {inst_id for inst_id, rows in capacities.items() if any(row[1] in codes for row in rows)}
        for inst_id in sorted(ids):
            if inst_id not in general:
                continue
            record = base_record(general[inst_id], category, evals)
            types = []
            for row in capacities.get(inst_id, []):
                if row[1] in codes:
                    types.append((row[1], clean(row[2])))
                    record["z"] += number(row[3])
            for row in staffing.get(inst_id, []):
                if row[1] in codes:
                    types.append((row[1], clean(row[2])))
                    merge_numeric(record, row)
            unique_types = list(dict.fromkeys(types))
            record["t"] = ",".join(code for code, _ in unique_types)
            record["tn"] = " · ".join(name for _, name in unique_types)
            output[category][inst_id] = record

    ids = {inst_id for inst_id, rows in capacities.items() if any(dementia(row[1]) for row in rows)}
    for inst_id in sorted(ids):
        if inst_id not in general:
            continue
        record = base_record(general[inst_id], "dementia", evals)
        types = []
        for row in capacities.get(inst_id, []):
            if dementia(row[1]):
                types.append((row[1], clean(row[2])))
                record["z"] += number(row[3])
        for row in staffing.get(inst_id, []):
            if dementia(row[1]):
                types.append((row[1], clean(row[2])))
                merge_numeric(record, row)
        unique_types = list(dict.fromkeys(types))
        record["t"] = ",".join(code for code, _ in unique_types)
        record["tn"] = " · ".join(name for _, name in unique_types)
        output["dementia"][inst_id] = record
    return {key: list(value.values()) for key, value in output.items()}


def build_hospitals():
    rows = []
    with HIRA_CSV.open("r", encoding="cp949", newline="") as handle:
        for row in csv.DictReader(handle):
            if clean(row.get("요양종별")) != "요양병원":
                continue
            encrypted_id = clean(row.get("암호화된요양기호"))
            # 원본에는 기준일 이후 폐업 안내 문구가 요양기호 칸에 들어간 5개 행이 있다.
            # 상세페이지 연결이 불가능한 해당 행은 통합지도에서 제외한다.
            if not encrypted_id.startswith("JD"):
                continue
            province = clean(row.get("시도명"))
            city = clean(row.get("시군구명"))
            rows.append({
                "i": encrypted_id, "n": clean(row.get("요양기관명")),
                "p": province, "c": city, "a": clean(row.get("도로명주소")),
                "d": iso_date(row.get("개설일자")), "t": "HOSP", "tn": "요양병원",
                "z": 0, "s": 0, "rn": 0, "na": 0, "pt": 0, "ot": 0, "cw": 0,
            })
    return rows


def compact_number(value):
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def write_chunks(category, rows):
    rows = [{key: compact_number(value) for key, value in row.items() if value not in (None, "")} for row in rows]
    rows.sort(key=lambda row: (PROVINCES.index(row.get("p", "")) if row.get("p", "") in PROVINCES else 99, row.get("c", ""), row.get("n", "")))
    chunks = []
    current = []
    current_size = 0
    for row in rows:
        encoded = json.dumps(row, ensure_ascii=False, separators=(",", ":"))
        if current and current_size + len(encoded) + 1 > TARGET_CHARS:
            chunks.append(current)
            current, current_size = [], 0
        current.append(row)
        current_size += len(encoded) + 1
    if current:
        chunks.append(current)

    files = []
    for index, chunk in enumerate(chunks, 1):
        filename = f"{category}-{index:02d}.js"
        payload = json.dumps(chunk, ensure_ascii=False, separators=(",", ":"))
        (OUT_DIR / filename).write_text(
            "window.NATIONAL_CARE_DATA=(window.NATIONAL_CARE_DATA||[]).concat(" + payload + ");\n",
            encoding="utf-8",
        )
        files.append("nationwide-care-data/" + filename)
    return files


def main():
    OUT_DIR.mkdir(exist_ok=True)
    for old in OUT_DIR.glob("*.js"):
        old.unlink()
    records = build_ltc_records()
    records["nursing-hospital"] = build_hospitals()
    manifest = {}
    for category, meta in CATEGORIES.items():
        rows = records[category]
        files = write_chunks(category, rows)
        eval_count = sum(1 for row in rows if row.get("g"))
        manifest[category] = {**meta, "count": len(rows), "evaluationCount": eval_count, "files": files}
    (ROOT / "nationwide-care-manifest.js").write_text(
        "window.NATIONAL_CARE_MANIFEST=" + json.dumps(manifest, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
