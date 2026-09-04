#!/usr/bin/env python3
"""국민건강보험공단 공개자료를 정적 JSON으로 정규화한다."""

# /** SOFTM-NHIS-STATIC START 날짜:20260903 : 브라우저의 공단 실시간 조회를 배포 전 정적 수집으로 대체하기 위한 공통 수집기 */
from __future__ import annotations

import argparse
import csv
import gzip
import hashlib
import html
import json
import os
import random
import re
import sys
import time
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import openpyxl
import requests


ROOT = Path(__file__).resolve().parent.parent
DATA_ROOT = ROOT / "data" / "nhis"
SOURCE_ROOT = ROOT / "source-data"
FIXTURE_ROOT = ROOT / "tests" / "fixtures" / "nhis"
DETAIL_BASE = "https://apis.data.go.kr/B550928/getLtcInsttDetailInfoService02"
SEARCH_BASE = "https://apis.data.go.kr/B550928/searchLtcInsttService02/getLtcInsttSeachList02"
PHOTO_PAGE = "https://www.longtermcare.or.kr/npbs/r/a/201/selectLtcoSrchDetail.web"
PHOTO_URL = "https://www.longtermcare.or.kr/npbs/attachfile/sendFileThumbnailTop.web?keyValue={key}"
MAX_DAILY_CALLS = 900_000  # SOFTM-NHIS-API-LIMIT 날짜:20260903 : 운영계정 일일 한도 100만 회 확인 후 10% 여유를 남기고 전체 상세 수집이 중단되지 않도록 상향
SCHEMA_VERSION = 1

# /** SOFTM-NHIS-OFFICIAL-PAGE START 날짜:20260903 : OpenAPI에 없는 기본정보·근속·CCTV도 공단 상세 화면과 같은 항목으로 보존 */
OFFICIAL_DETAIL_TABS = (11, 14, 19)
OFFICIAL_TAB_LABELS = {11: "기본정보", 14: "인력현황", 19: "CCTV현황"}
DETAIL_PROFILE = "official-page-tabs-11-14-19-v1"
# /** SOFTM-NHIS-OFFICIAL-PAGE END */

PROVINCE_CODES = ["11", "26", "27", "28", "29", "30", "31", "36", "41", "43", "44", "46", "47", "48", "50", "51", "52"]  # SOFTM-NHIS-REGION-CODE 날짜:20260903 : 특별자치도 전환 뒤 강원·전북 기관이 목록 수집에서 누락되지 않도록 현행 시도코드 사용
CATEGORY_CODES = {
    "facility": {"A01", "A02", "A03", "A04", "A05"},
    "daycare": {"B03", "C03"},
    "home-care": {"B01", "C01"},
    "home-nursing": {"B05", "C05"},
    "home-bath": {"B02", "C02"},
    "short-stay": {"B04", "C04"},
}
DETAIL_OPERATIONS = {
    "general": "getGeneralSttusDetailInfoItem02",
    "capacity": "getAceptncNmprDetailInfoItem02",
    "staff": "getStaffSttusDetailInfoItem02",
    "facilities": "getInsttSttusDetailInfoItem02",
    "nonCovered": "getNonBenefitSttusDetailInfoList02",
    "programs": "getProgramSttusDetailInfoList02",
    "agreements": "getConvInsttDetailInfoList02",
    "equipment": "getWlfareToolDetailInfoList02",
    "other": "getInsttEtcDetailInfoItem02",
}
LIST_OPERATIONS = {"nonCovered", "programs", "agreements", "equipment"}
TYPE_LABELS = {
    "A01": "노인요양시설", "A02": "노인요양공동생활가정", "A03": "양로시설", "A04": "노인공동생활가정", "A05": "노인복지주택",
    "B01": "방문요양", "B02": "방문목욕", "B03": "주야간보호", "B04": "단기보호", "B05": "방문간호",
    "C01": "방문요양", "C02": "방문목욕", "C03": "주야간보호", "C04": "단기보호", "C05": "방문간호",
}
FILE_SOURCES = {
    "facility": {"id": "15124763", "extension": "xlsx", "prefix": "nhis-longtermcare-"},
    "evaluation": {"id": "15104801", "extension": "csv", "prefix": "nhis_longtermcare_evaluations_"},
}


def clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").split())


def digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def iso_date(value: Any) -> str:
    text = digits(value)
    return f"{text[:4]}-{text[4:6]}-{text[6:8]}" if len(text) >= 8 else ""


def number(value: Any) -> int | float | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        result = float(text)
        return int(result) if result.is_integer() else result
    except ValueError:
        return None


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def json_bytes(value: Any) -> bytes:
    return (json.dumps(value, ensure_ascii=False, separators=(",", ":"), sort_keys=True) + "\n").encode("utf-8")


def write_json(path: Path, value: Any) -> bool:
    payload = json_bytes(value)
    if path.exists() and path.read_bytes() == payload:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_bytes(payload)
    temporary.replace(path)
    return True


def load_json(path: Path, default: Any) -> Any:
    try:
        if path.suffix == ".gz":
            with gzip.open(path, "rt", encoding="utf-8") as stream:
                return json.load(stream)
        return json.loads(path.read_text(encoding="utf-8"))
    except (FileNotFoundError, OSError, json.JSONDecodeError):
        return default


def write_gzip_json(path: Path, payload: Any) -> None:
    """상세 정적 파일을 재현 가능한 gzip으로 원자 저장한다."""
    # /** SOFTM-NHIS-GZIP START 날짜:20260903 : 전체 상세가 GitHub Pages 용량 한도를 넘지 않도록 기관별 JSON을 직접 압축 저장 */
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    encoded = (json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n").encode("utf-8")
    with temporary.open("wb") as raw_stream:
        with gzip.GzipFile(filename="", mode="wb", compresslevel=9, fileobj=raw_stream, mtime=0) as gzip_stream:
            gzip_stream.write(encoded)
    os.replace(temporary, path)
    # /** SOFTM-NHIS-GZIP END */


def load_local_env() -> None:
    path = ROOT / ".env.local"
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line or line.lstrip().startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


def refresh_source_file(source: dict[str, str]) -> Path:
    """공공데이터포털이 현재 가리키는 원문 파일을 확인하고 새 버전만 원자적으로 보관한다."""
    public_id, extension = source["id"], source["extension"]
    page_url = f"https://www.data.go.kr/data/{public_id}/fileData.do"
    session = requests.Session()
    session.headers.update({"User-Agent": "NationwideCareStaticCollector/1.0 (+https://softm.github.io/)"})
    page = session.get(page_url, timeout=(8, 30))
    page.raise_for_status()
    detail_match = re.search(r'id="publicDataDetailPk"[^>]+value="([^"]+)"', page.text)
    call_match = re.search(rf"fn_fileDataDown\('{public_id}',\s*'([^']+)',\s*'[^']*',\s*'(\d+)',\s*'(\d+)'\)", page.text)
    if not detail_match or not call_match:
        raise RuntimeError(f"공공데이터 {public_id} 다운로드 메타데이터를 찾을 수 없습니다.")
    detail_id, file_detail_sn = detail_match.group(1), call_match.group(2)
    metadata_response = session.post(
        "https://www.data.go.kr/tcs/dss/selectFileDataDownload.do",
        data={"publicDataDetailPk": detail_id, "publicDataPk": public_id, "atchFileId": "", "fileDetailSn": file_detail_sn, "publicDataTyCode": "PR0051"},
        headers={"X-Requested-With": "XMLHttpRequest", "Referer": page_url}, timeout=(8, 30),
    )
    metadata_response.raise_for_status()
    metadata = metadata_response.json()
    if not metadata.get("status") or not metadata.get("atchFileId"):
        raise RuntimeError(f"공공데이터 {public_id} 다운로드 정보 확인에 실패했습니다.")
    download = session.get(
        "https://www.data.go.kr/cmm/cmm/fileDownload.do",
        params={"atchFileId": metadata["atchFileId"], "fileDetailSn": metadata.get("fileDetailSn", file_detail_sn), "insertDataPrcus": "N"},
        headers={"Referer": page_url}, timeout=(8, 90),
    )
    download.raise_for_status()
    payload = download.content
    if extension == "xlsx" and not payload.startswith(b"PK"):
        raise RuntimeError(f"공공데이터 {public_id} 응답이 XLSX가 아닙니다.")
    if extension == "csv" and (not payload or payload.lstrip().startswith((b"<", b"{"))):
        raise RuntimeError(f"공공데이터 {public_id} 응답이 CSV가 아닙니다.")
    info = metadata.get("dataSetFileDetailInfo") or {}
    date_match = re.search(r"(20\d{6})", clean(info.get("dataNm")) + page.text)
    if not date_match:
        raise RuntimeError(f"공공데이터 {public_id} 기준일을 찾을 수 없습니다.")
    path = SOURCE_ROOT / f"{source['prefix']}{date_match.group(1)}.{extension}"
    if path.exists() and path.read_bytes() == payload:
        return path
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_bytes(payload)
    temporary.replace(path)
    return path


def latest_source(pattern: str, label: str) -> Path:
    paths = sorted(SOURCE_ROOT.glob(pattern))
    if not paths:
        raise FileNotFoundError(f"{label} 원본 파일이 없습니다: source-data/{pattern}")
    return paths[-1]


def category_for(code: str) -> str:
    for category, codes in CATEGORY_CODES.items():
        if code in codes:
            return category
    return "dementia" if code.startswith(("G", "H", "I", "M")) or code == "S41" else "other"


class ApiBudgetExceeded(RuntimeError):
    pass


class DataGoClient:
    def __init__(self, service_key: str, max_calls: int = MAX_DAILY_CALLS, fixture_dir: Path | None = None):
        self.service_key = service_key
        self.max_calls = max_calls
        self.fixture_dir = fixture_dir
        self.calls = 0
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "NationwideCareStaticCollector/1.0 (+https://softm.github.io/)"})

    def request_xml(self, url: str, params: dict[str, Any], fixture_name: str = "") -> ET.Element:
        if not self.service_key:
            fixture = self.fixture_dir / f"{fixture_name}.xml" if self.fixture_dir and fixture_name else None
            if fixture and fixture.exists():
                return ET.fromstring(fixture.read_text(encoding="utf-8"))
            raise RuntimeError("DATA_GO_KR_SERVICE_KEY가 없고 사용할 XML fixture도 없습니다.")
        if self.calls >= self.max_calls:
            raise ApiBudgetExceeded(f"공공데이터 일일 호출 상한 {self.max_calls:,}회에 도달했습니다.")
        self.calls += 1
        query = {"serviceKey": self.service_key, **params}
        last_error: Exception | None = None
        for attempt in range(5):
            try:
                response = self.session.get(url, params=query, timeout=(8, 25))
                # /** SOFTM-NHIS-API-LIMIT START 날짜:20260903 : 실제 일일 호출 제한에 도달하면 불필요한 재시도와 실패 문서 생성을 즉시 중단 */
                if response.status_code == 429:
                    raise ApiBudgetExceeded("공공데이터 API가 HTTP 429 호출 제한을 반환했습니다.")
                if response.status_code in {500, 502, 503, 504}:
                    raise requests.HTTPError(f"HTTP {response.status_code}", response=response)
                # /** SOFTM-NHIS-API-LIMIT END */
                response.raise_for_status()
                root = ET.fromstring(response.content)
                code = root.findtext(".//resultCode") or "00"
                if code not in {"00", "0", "NORMAL_SERVICE"}:
                    message = root.findtext(".//resultMsg") or "공공데이터 API 오류"
                    raise RuntimeError(f"API {code}: {message}")
                return root
            except ApiBudgetExceeded:  # SOFTM-NHIS-API-LIMIT 날짜:20260903 : HTTP 429를 일반 통신 오류로 다시 잡지 않고 전체 수집 체크포인트까지 즉시 전달
                raise
            except (requests.RequestException, ET.ParseError, RuntimeError) as error:
                last_error = error
                if attempt == 4:
                    break
                time.sleep(min(10, (2 ** attempt) + random.uniform(0.15, 0.85)))
        raise RuntimeError(f"공공데이터 API 요청 실패: {last_error}")


def xml_items(root: ET.Element) -> list[dict[str, str]]:
    output = []
    for item in root.findall(".//item"):
        output.append({child.tag: clean(child.text) for child in item if clean(child.text)})
    return output


def first_item(root: ET.Element) -> dict[str, str]:
    rows = xml_items(root)
    return rows[0] if rows else {}


def fetch_all_pages(client: DataGoClient, url: str, params: dict[str, Any], fixture_name: str = "") -> list[dict[str, str]]:
    page = 1
    rows: list[dict[str, str]] = []
    while True:
        root = client.request_xml(url, {**params, "pageNo": page, "numOfRows": 1000}, fixture_name if page == 1 else "")
        batch = xml_items(root)
        rows.extend(batch)
        total = int(root.findtext(".//totalCount") or len(rows))
        if not batch or len(rows) >= total:
            return rows
        page += 1


def build_xlsx_catalog() -> list[dict[str, Any]]:
    path = latest_source("nhis-longtermcare-*.xlsx", "시설별 현황")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    general = {digits(row[0]): row for row in workbook["일반현황"].iter_rows(min_row=2, values_only=True) if digits(row[0])}
    grouped: dict[str, dict[str, Any]] = {}
    for row in workbook["입소인원"].iter_rows(min_row=2, values_only=True):
        institution_id, service_code = digits(row[0]), clean(row[1]).upper()
        if not institution_id or institution_id not in general or not service_code:
            continue
        values = general[institution_id]
        entry = grouped.setdefault(institution_id, {
            "id": institution_id, "name": clean(values[1]), "postalCode": clean(values[2]),
            "regionCodes": {"province": clean(values[3]), "city": clean(values[4]), "district": clean(values[5])},
            "region": clean(values[6]), "address": clean(values[9]), "phone": "",
            "designationDate": iso_date(values[7]), "installationDate": iso_date(values[8]), "services": [],
        })
        if not any(service["code"] == service_code for service in entry["services"]):
            entry["services"].append({"code": service_code, "name": clean(row[2]) or TYPE_LABELS.get(service_code, service_code), "category": category_for(service_code), "capacity": number(row[3])})
    workbook.close()
    for entry in grouped.values():
        entry["services"].sort(key=lambda item: item["code"])
    return sorted(grouped.values(), key=lambda item: item["id"])


def merge_live_catalog(client: DataGoClient, catalog: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], set[str]]:
    indexed = {item["id"]: item for item in catalog}
    seen: set[str] = set()
    for province_code in PROVINCE_CODES:
        rows = fetch_all_pages(client, SEARCH_BASE, {"siDoCd": province_code})
        for row in rows:
            institution_id = digits(row.get("longTermAdminSym"))
            service_code = clean(row.get("adminPttnCd")).upper()
            if not institution_id:
                continue
            seen.add(institution_id)
            item = indexed.setdefault(institution_id, {
                "id": institution_id, "name": clean(row.get("adminNm")), "postalCode": "", "regionCodes": {},
                "region": "", "address": "", "phone": "", "designationDate": iso_date(row.get("longTermPeribRgtDt")),
                "installationDate": "", "services": [],
            })
            if clean(row.get("adminNm")):
                item["name"] = clean(row.get("adminNm"))
            item["installationDate"] = iso_date(row.get("stpRptDt")) or item.get("installationDate", "")
            if service_code and not any(service["code"] == service_code for service in item["services"]):
                item["services"].append({"code": service_code, "name": TYPE_LABELS.get(service_code, service_code), "category": category_for(service_code), "capacity": None})
    merged = sorted(indexed.values(), key=lambda item: item["id"])
    for item in merged:
        item["services"].sort(key=lambda service: service["code"])
    return merged, seen


def catalog_hash(item: dict[str, Any]) -> str:
    watched = {key: item.get(key) for key in ("name", "address", "phone", "designationDate", "installationDate", "closedDate", "services")}
    return hashlib.sha256(json_bytes(watched)).hexdigest()[:16]


def catalog_changes(old_rows: list[dict[str, Any]], new_rows: list[dict[str, Any]], live_seen: set[str]) -> tuple[dict[str, Any], set[str]]:
    old = {item["id"]: item for item in old_rows}
    new = {item["id"]: item for item in new_rows}
    added = sorted(set(new) - set(old))
    removed = sorted((set(old) - live_seen) if live_seen else set())
    changed = []
    changed_ids = set(added)
    for institution_id in sorted(set(old) & set(new)):
        fields = [key for key in ("name", "address", "phone", "designationDate", "installationDate", "services") if old[institution_id].get(key) != new[institution_id].get(key)]
        if fields:
            changed.append({"id": institution_id, "fields": fields})
            changed_ids.add(institution_id)
    return {"added": added, "removed": removed, "changed": changed}, changed_ids


def detail_params(institution_id: str, service_code: str, section: str) -> dict[str, Any]:
    params: dict[str, Any] = {"longTermAdminSym": institution_id, "adminPttnCd": service_code}
    if section in LIST_OPERATIONS:
        params.update({"pageNo": 1, "numOfRows": 100})
    return params


def normalise_detail_section(section: str, rows: list[dict[str, str]]) -> Any:
    if not rows:
        return None
    if section == "general":
        row = rows[0]
        phone = "-".join(part for part in (clean(row.get("locTelNo_1")), clean(row.get("locTelNo_2")), clean(row.get("locTelNo_3"))) if part)
        road_number = "-".join(part for part in (clean(row.get("gunmulMlno")), clean(row.get("gunmulSlno"))) if part and part != "0")
        return {
            "name": clean(row.get("adminNm")), "postalCode": clean(row.get("hmPostNo")), "phone": phone,
            "regionCodes": {"province": clean(row.get("siDoCd")), "city": clean(row.get("siGunGuCd")), "district": clean(row.get("HDongCd") or row.get("BDongCd")), "village": clean(row.get("riCd"))},
            "detailAddress": clean(row.get("detailAddr")), "roadNameCode": clean(row.get("roadNmCd")), "buildingNumber": road_number,
            "floor": clean(row.get("fl")), "designationDate": iso_date(row.get("longTermPeribRgtDt")), "installationDate": iso_date(row.get("stpRptDt")),
        }
    if section == "capacity":
        row = rows[0]
        return {"capacity": number(row.get("totPer")), "current": {"male": number(row.get("maNowPer")), "female": number(row.get("fmNowPer"))}, "reserved": {"male": number(row.get("maRsvPer")), "female": number(row.get("fmRsvPer"))}}
    if section == "staff":
        row = rows[0]
        mapping = {"equipmentLong": "equipLong", "director": "hdOfce", "socialWorker": "socWel", "doctor": "chrgDoc", "contractDoctor": "chargeDoc", "nurse": "nur", "nursingAssistant": "nurArticle", "dentalHygienist": "dent", "physicalTherapist": "physicalMTret", "occupationalTherapist": "wrkMTret", "careWorker": "recuProt_1", "careWorkerSecondary": "recuProt_2", "careWorkerDeferred": "recuProtDelay", "officeWorker": "ofceEmp", "dietitian": "nut", "cook": "cook", "hygieneWorker": "hygiPrsn", "manager": "mgmtPrsn", "supportWorker": "suppPrsn", "other": "etcPer"}
        return {key: number(row.get(source)) for key, source in mapping.items()}
    if section == "facilities":
        row = rows[0]
        mapping = {"singleRooms": "prsnRoomReal1", "doubleRooms": "prsnRoomReal2", "tripleRooms": "prsnRoomReal3", "fourPlusRooms": "prsnRoomReal4", "specialBathRooms": "spcAcupRoomReal", "offices": "ofce", "medicalRooms": "medRoomReal", "trainingRooms": "funcTrnRoomReal", "programRooms": "pgmRoomReal", "cafeterias": "crmnyPrst", "bathRooms": "batRoom", "washRooms": "taxRoom", "washBasins": "taxPageLong"}
        return {key: number(row.get(source)) for key, source in mapping.items()}
    if section == "other":
        row = rows[0]
        return {"homepage": clean(row.get("hmpgAddr")), "transport": clean(row.get("tfMth")), "parking": clean(row.get("pkngEquip"))}
    if section == "nonCovered":
        return [{"kind": clean(row.get("nonpayKind")), "basis": clean(row.get("prodBase")), "amount": number(row.get("nonpayTgtAmt")), "updatedDate": iso_date(row.get("uptDt"))} for row in rows]
    if section == "programs":
        return [{"type": clean(row.get("pgmType")), "name": clean(row.get("pgmNm")), "targetCount": number(row.get("tgtNop")), "cycle": clean(row.get("cyclTm")), "place": clean(row.get("runPlc"))} for row in rows]
    if section == "agreements":
        return [{"name": clean(row.get("yoyangNm")), "from": iso_date(row.get("adptFrDt")), "to": iso_date(row.get("adptToDt"))} for row in rows]
    if section == "equipment":
        return [{"reportDescription": clean(row.get("witemRptDesc")), "name": clean(row.get("itemName")), "manufacturer": clean(row.get("mnfCo")), "model": clean(row.get("modelNm")), "usage": clean(row.get("usage")), "note": clean(row.get("rmk"))} for row in rows]
    return rows


# /** SOFTM-NHIS-OFFICIAL-PAGE START 날짜:20260903 : 공단 상세 화면에서 OpenAPI 비제공 표와 기본 항목을 구조화 */
class OfficialDetailTableParser(HTMLParser):
    """공단 상세 탭의 표를 표시 순서와 셀 병합 정보까지 보존한다."""

    def __init__(self):
        super().__init__()
        self.tables: list[dict[str, Any]] = []
        self.table: dict[str, Any] | None = None
        self.row: list[dict[str, Any]] | None = None
        self.cell: dict[str, Any] | None = None
        self.cell_parts: list[str] = []
        self.in_caption = False
        self.caption_parts: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        values = dict(attrs)
        if tag == "table" and self.table is None:
            self.table = {"caption": "", "rows": []}
        elif tag == "caption" and self.table is not None:
            self.in_caption, self.caption_parts = True, []
        elif tag == "tr" and self.table is not None:
            self.row = []
        elif tag in {"th", "td"} and self.row is not None:
            self.cell = {
                "header": tag == "th",
                "colspan": max(1, int(values.get("colspan") or 1)),
                "rowspan": max(1, int(values.get("rowspan") or 1)),
            }
            self.cell_parts = []

    def handle_data(self, data: str) -> None:
        if self.cell is not None:
            self.cell_parts.append(data)
        elif self.in_caption:
            self.caption_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"th", "td"} and self.cell is not None and self.row is not None:
            self.cell["text"] = clean("".join(self.cell_parts))
            self.row.append(self.cell)
            self.cell, self.cell_parts = None, []
        elif tag == "tr" and self.row is not None and self.table is not None:
            if self.row:
                self.table["rows"].append({"cells": self.row})
            self.row = None
        elif tag == "caption" and self.in_caption and self.table is not None:
            self.table["caption"] = clean("".join(self.caption_parts))
            self.in_caption, self.caption_parts = False, []
        elif tag == "table" and self.table is not None:
            if self.table["rows"]:
                self.tables.append(self.table)
            self.table = None


def is_facility_service(service_code: str) -> bool:
    return service_code in {"A01", "A02", "A03", "A04", "A05", "B03", "B04", "C03", "C04", "S41"} or bool(re.fullmatch(r"[GHIM][3-9][1-9]", service_code))


def parse_official_detail_tab(page_html: str, tab: int, service_code: str) -> dict[str, Any]:
    parser = OfficialDetailTableParser()
    parser.feed(page_html)
    tables = parser.tables
    if tab == 14 and len(tables) >= 4:
        tables = tables[:2] if is_facility_service(service_code) else tables[2:4]
    changed_match = re.search(r"최종변경일\s*:\s*(20\d{2}[.\-/]\d{2}[.\-/]\d{2})", page_html)
    result: dict[str, Any] = {
        "tab": tab, "label": OFFICIAL_TAB_LABELS[tab], "collectedAt": now_iso(),
        "lastModifiedDate": iso_date(changed_match.group(1)) if changed_match else "", "tables": tables,
    }
    if tab == 11:
        fields: dict[str, str] = {}
        for row in (tables[0].get("rows", []) if tables else []):
            cells = row.get("cells", [])
            if len(cells) >= 2 and clean(cells[0].get("text")):
                fields[clean(cells[0]["text"])] = clean(" ".join(cell.get("text", "") for cell in cells[1:]))
        result["fields"] = fields
    return result


def request_official_detail_tab(session: requests.Session, institution_id: str, service_code: str, tab: int) -> str:
    params = {
        "aTab": str(tab), "adminPttnCd": service_code, "ltcAdminSym": institution_id,
        "paymtVltClsfcTypeCd": "", "paymtVltClsfcTypeCdSusi": "", "paymtVltMgmtNo": "",
        "paymtVltMgmtNo2": "", "paymtVltMgmtNoOld": "", "showVlt": "Y", "vltMgmtYyyy": "",
    }
    last_error: Exception | None = None
    for attempt in range(4):
        try:
            response = session.get(PHOTO_PAGE, params=params, timeout=(8, 30))
            response.raise_for_status()
            if institution_id not in response.text or f'id="tab_{tab}"' not in response.text:
                raise RuntimeError(f"공단 상세 탭 {tab} 응답 형식이 올바르지 않습니다.")
            return response.text
        except (requests.RequestException, RuntimeError) as error:
            last_error = error
            if attempt < 3:
                time.sleep(min(6, (2 ** attempt) + random.uniform(0.1, 0.5)))
    raise RuntimeError(f"공단 상세 탭 {tab} 요청 실패: {last_error}")


def collect_official_detail_page(institution: dict[str, Any], service_code: str) -> tuple[dict[str, Any], list[dict[str, str]]]:
    session = requests.Session()
    session.headers.update({"User-Agent": "NationwideCareStaticCollector/1.0", "Accept-Language": "ko-KR,ko;q=0.9"})
    tabs: dict[str, Any] = {}
    failures: list[dict[str, str]] = []
    for tab in OFFICIAL_DETAIL_TABS:
        try:
            parsed = parse_official_detail_tab(request_official_detail_tab(session, institution["id"], service_code, tab), tab, service_code)
            if not parsed.get("tables") or (tab == 11 and not parsed.get("fields")):
                raise RuntimeError(f"공단 상세 탭 {tab}에서 표시 항목을 찾지 못했습니다.")
            tabs[str(tab)] = parsed
        except Exception as error:
            failures.append({"section": f"officialPage:{tab}", "message": str(error)[:300]})
        time.sleep(random.uniform(0.08, 0.18))
    return {
        "source": "longtermcare.or.kr:selectLtcoSrchDetail", "collectedAt": now_iso(),
        "institutionId": institution["id"], "serviceCode": service_code, "tabs": tabs,
    }, failures
# /** SOFTM-NHIS-OFFICIAL-PAGE END */


def fetch_detail(client: DataGoClient, institution: dict[str, Any], service_code: str, include_official_page: bool = True) -> dict[str, Any]:
    institution_id = institution["id"]
    sections: dict[str, Any] = {}
    failures = []
    for section, operation in DETAIL_OPERATIONS.items():
        try:
            root = client.request_xml(f"{DETAIL_BASE}/{operation}", detail_params(institution_id, service_code, section), f"{operation}-{institution_id}-{service_code}")
            sections[section] = normalise_detail_section(section, xml_items(root))
        except ApiBudgetExceeded:
            raise
        except Exception as error:
            failures.append({"section": section, "message": str(error)[:300]})
    # /** SOFTM-NHIS-OFFICIAL-PAGE START 날짜:20260903 : OpenAPI 누락 항목은 공단 공개 상세 페이지 탭에서 보완 */
    official_page, official_failures = collect_official_detail_page(institution, service_code) if include_official_page else ({"tabs": {}}, [])
    failures.extend(official_failures)
    if official_page.get("tabs", {}).get("19"):
        sections["cctv"] = official_page["tabs"]["19"].get("tables", [])
    unavailable = ["photos"] + ([] if sections.get("cctv") is not None else ["cctv"])
    # /** SOFTM-NHIS-OFFICIAL-PAGE END */
    return {
        "serviceCode": service_code, "collectedAt": now_iso(), "sections": sections,
        "availableSections": [key for key, value in sections.items() if value not in (None, [], {})],
        "unavailableSections": unavailable, "failures": failures, "officialPage": official_page,
    }


def merge_detail_document(existing: dict[str, Any], institution: dict[str, Any], detail: dict[str, Any], evaluations: dict[str, Any]) -> dict[str, Any]:
    service_details = existing.get("serviceDetails", {}) if existing.get("id") == institution["id"] else {}
    # /** SOFTM-NHIS-OFFICIAL-PAGE START 날짜:20260903 : 일시적인 공단 탭 장애가 이전에 성공한 화면 스냅샷을 지우지 않도록 병합 */
    previous_detail = service_details.get(detail["serviceCode"], {})
    previous_tabs = previous_detail.get("officialPage", {}).get("tabs", {})
    official_page = detail.get("officialPage", {})
    official_page["tabs"] = {**previous_tabs, **official_page.get("tabs", {})}
    detail["officialPage"] = official_page
    if "19" in official_page["tabs"]:
        detail["sections"]["cctv"] = official_page["tabs"]["19"].get("tables", [])
        detail["unavailableSections"] = [item for item in detail["unavailableSections"] if item != "cctv"]
    # /** SOFTM-NHIS-OFFICIAL-PAGE END */
    service_details[detail["serviceCode"]] = detail
    sections = detail["sections"]
    fetched_at = now_iso()
    return {
        "schemaVersion": SCHEMA_VERSION, "institutionId": institution["id"], "id": institution["id"],
        "institutionTypeCode": detail["serviceCode"], "name": institution.get("name", ""),
        "fetchedAt": fetched_at, "collectedAt": fetched_at,
        "sources": {"catalog": "data-go-kr:15059029", "facilityFile": "data-go-kr:15124763", "detail": "data-go-kr:15058856", "evaluation": "data-go-kr:15104801", "officialPage": "longtermcare.or.kr:selectLtcoSrchDetail"}, # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : 화면 고유 항목의 출처를 OpenAPI와 구분
        "basic": {**institution, **(sections.get("general") or {})}, "services": institution.get("services", []),
        "capacity": sections.get("capacity"), "staff": sections.get("staff"), "facility": sections.get("facilities"),
        "nonCovered": sections.get("nonCovered") or [], "programs": sections.get("programs") or [],
        "agreements": sections.get("agreements") or [], "cctv": sections.get("cctv"), "evaluation": evaluations.get(institution["id"], []),
        "officialPage": detail.get("officialPage", {}), # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : 대표 급여의 공단 화면 스냅샷도 기존 호환 필드와 함께 제공
        "availableSections": detail["availableSections"], "unavailableSections": detail["unavailableSections"],
        "errors": detail["failures"], "serviceDetails": service_details,
    }


def evaluation_rows() -> dict[str, Any]:
    path = latest_source("nhis_longtermcare_evaluations_*.csv", "평가 결과")
    latest: dict[str, dict[str, tuple[str, dict[str, Any]]]] = defaultdict(dict)
    with path.open(encoding="cp949", newline="") as handle:
        for row in csv.DictReader(handle):
            institution_id = digits(row.get("장기요양기관기호"))
            service = clean(row.get("급여종류"))
            date = clean(row.get("평가일자"))
            previous = latest[institution_id].get(service)
            if previous and previous[0] >= date:
                continue
            latest[institution_id][service] = (date, {
                "service": service, "evaluation": clean(row.get("평가구분")), "date": date,
                "grade": clean(row.get("평가등급")), "score": number(row.get("평가총점")),
                "operation": number(row.get("기관운영")), "safety": number(row.get("환경및안전")),
                "rights": number(row.get("수급자권리보장")), "process": number(row.get("급여제공과정")),
                "result": number(row.get("급여제공결과")), "operation2025": number(row.get("기관운영(2025)")),
                "respect2025": number(row.get("수급자존중(2025)")), "service2025": number(row.get("서비스제공(2025)")),
                "result2025": number(row.get("서비스결과(2025)")),
            })
    return {institution_id: [value for _, value in sorted(items.values(), key=lambda pair: pair[0], reverse=True)] for institution_id, items in sorted(latest.items())}


class PhotoListParser(HTMLParser):
    """공단 사진 목록의 공개 메타데이터만 읽고 본문·바이너리는 저장하지 않는다."""

    def __init__(self):
        super().__init__()
        self.in_item = False
        self.item_depth = 0
        self.current: dict[str, str] = {}
        self.link = ""
        self.text_parts: list[str] = []
        self.items: list[dict[str, str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        values = dict(attrs)
        if tag == "li" and not self.in_item:
            self.in_item, self.item_depth, self.current = True, 1, {}
            return
        if not self.in_item:
            return
        if tag == "li":
            self.item_depth += 1
        if tag == "a":
            self.link = values.get("href") or ""
            self.text_parts = []
            if self.link and "selectBlbdArtiDtl.web" in self.link:
                self.current.setdefault("detailUrl", urljoin(PHOTO_PAGE, self.link))
        if tag == "img":
            source = values.get("src") or ""
            match = re.search(r"sendFileThumbnailTop\.web\?keyValue=([A-Za-z0-9_-]+)", source, re.I)
            if match:
                key = match.group(1)
                self.current.update({"key": key, "thumbnailUrl": PHOTO_URL.format(key=key), "url": PHOTO_URL.format(key=key)})

    def handle_data(self, data: str) -> None:
        if self.in_item and self.link:
            self.text_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        if not self.in_item:
            return
        if tag == "a" and self.link:
            value = clean(" ".join(self.text_parts))
            if value:
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
                    self.current["date"] = value
                elif self.link and "selectBlbdArtiDtl.web" in self.link and not self.current.get("title"):
                    self.current["title"] = value
            self.link, self.text_parts = "", []
        if tag == "li":
            self.item_depth -= 1
            if self.item_depth == 0:
                if self.current.get("key"):
                    self.items.append(self.current)
                self.in_item, self.current = False, {}


def parse_photos(page_html: str, max_photos: int, page_number: int = 1) -> list[dict[str, str | int]]:
    parser = PhotoListParser()
    parser.feed(page_html)
    photos: list[dict[str, str | int]] = []
    seen: set[str] = set()
    for item in parser.items:
        key = item["key"]
        if key in seen:
            continue
        seen.add(key)
        photos.append({
            "key": key, "title": item.get("title", ""), "date": item.get("date", ""),
            "category": "", "categoryContext": "전체", "thumbnailUrl": item["thumbnailUrl"], "url": item["url"],
            "detailUrl": item.get("detailUrl", ""), "page": page_number, "alt": item.get("title") or f"기관 등록사진 {len(photos) + 1}",
        })
        if len(photos) >= max_photos:
            break
    return photos


# /** SOFTM-PHOTO-FULL-TITLE START 날짜:20260904 : 공단 목록에서 이미 잘린 제목은 사진 게시글의 원래 제목으로 보완 */
class PhotoTitleParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.depth = 0
        self.parts: list[str] = []

    def handle_starttag(self, tag, attrs):
        if self.depth:
            if tag == "br":
                self.parts.append(" ")
            if tag not in {"br", "img", "hr", "input", "meta", "link"}:
                self.depth += 1
        elif tag == "span" and "tbl_tit" in (dict(attrs).get("class") or "").split():
            self.depth = 1

    def handle_endtag(self, tag):
        if self.depth and tag not in {"br", "img", "hr", "input", "meta", "link"}:
            self.depth -= 1

    def handle_data(self, value):
        if self.depth:
            self.parts.append(value)


def needs_photo_title(photo):
    return photo.get("titleSource") != "detail" and bool(re.search(r"(?:\.{3}|…)\s*$", photo.get("title") or photo.get("alt") or ""))


def fill_photo_title(session, photo):
    if not needs_photo_title(photo):
        return
    response = session.get(photo["detailUrl"], timeout=(8, 25))
    response.raise_for_status()
    parser = PhotoTitleParser()
    parser.feed(response.text)
    title = clean("".join(parser.parts))
    if not title:
        raise ValueError("사진 게시글의 전체 제목을 찾지 못했습니다.")
    photo.update(title=title, alt=title, titleSource="detail")
    photo.pop("titleStatus", None)
# /** SOFTM-PHOTO-FULL-TITLE END */


def collect_photos(institution: dict[str, Any], service_code: str, max_photos: int, mode: str) -> dict[str, Any]:
    params = {"ltcAdminSym": institution["id"], "adminPttnCd": service_code, "aTab": "18"}
    session = requests.Session()
    session.headers.update({"User-Agent": "NationwideCareStaticCollector/1.0", "Accept-Language": "ko-KR,ko;q=0.9"})
    response = session.get(PHOTO_PAGE, params=params, timeout=(8, 25))
    response.raise_for_status()
    photos = parse_photos(response.text, max_photos, 1)
    total_match = re.search(r"totalPageCount:\s*(\d+)", response.text)
    total_pages = int(total_match.group(1)) if total_match else 1
    page_number = 2
    while len(photos) < max_photos and page_number <= total_pages:
        page_response = session.post(PHOTO_PAGE, data={
            **params, "blbdId": "2203", "rcdCnt": "12", "artiCtgryVl": "", "pgmId": "npra201t18s",
            "pageInfo.pageSize": "10", "pageInfo.recordCountPerPage": "12", "pageInfo.currentPageNo": str(page_number),
        }, timeout=(8, 25))
        page_response.raise_for_status()
        existing_keys = {photo["key"] for photo in photos}
        photos.extend(photo for photo in parse_photos(page_response.text, max_photos - len(photos), page_number) if photo["key"] not in existing_keys)
        page_number += 1
        time.sleep(random.uniform(0.15, 0.45))
    # /** SOFTM-PHOTO-FULL-TITLE START 날짜:20260904 : 이후 사진 수집에서도 줄임표가 포함된 목록 제목을 그대로 저장하지 않도록 보완 */
    for photo in photos:
        try:
            fill_photo_title(session, photo)
        except ValueError:
            photo["titleStatus"] = "unavailable"
    # /** SOFTM-PHOTO-FULL-TITLE END */
    checked_at = now_iso()
    return {"schemaVersion": SCHEMA_VERSION, "institutionId": institution["id"], "id": institution["id"], "institutionTypeCode": service_code, "serviceCode": service_code, "checkedAt": checked_at, "collectedAt": checked_at, "mode": mode, "categoryContext": "전체", "photos": photos, "count": len(photos), "source": "longtermcare.or.kr public photo page"}


def select_institutions(args: argparse.Namespace, catalog: list[dict[str, Any]], changed_ids: set[str]) -> list[dict[str, Any]]:
    indexed = {item["id"]: item for item in catalog}
    if args.mode in {"institution", "fixture"}:
        ids = [digits(value) for value in args.institution]
    elif args.mode == "incremental":
        ids = sorted(changed_ids)
    elif args.mode == "retry":
        detail_failures = load_json(DATA_ROOT / "failures" / "details.json", {"items": []})
        photo_failures = load_json(DATA_ROOT / "failures" / "photos.json", {"items": []})
        ids = [digits(item.get("id")) for item in detail_failures.get("items", []) + photo_failures.get("items", [])]
    else:
        ids = [item["id"] for item in catalog if int(hashlib.sha256(item["id"].encode()).hexdigest(), 16) % args.shard_count == args.shard_index]
    selected = [indexed[institution_id] for institution_id in dict.fromkeys(ids) if institution_id in indexed]
    return selected[:args.limit_institutions] if args.limit_institutions else selected


def preferred_service(institution: dict[str, Any], requested_type: str) -> str:
    codes = [service["code"] for service in institution.get("services", [])]
    return requested_type if requested_type and requested_type in codes else (codes[0] if codes else requested_type)


def resume_shard(args: argparse.Namespace, targets: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any], Path | None]:
    if args.mode not in {"rotation", "full"}:
        return targets, {}, None
    path = DATA_ROOT / "checkpoints" / f"{args.mode}-{args.shard_index:02d}.json"
    state = {} if args.force else load_json(path, {})
    compatible = state.get("mode") == args.mode and state.get("shard") == {"index": args.shard_index, "count": args.shard_count} and state.get("detailProfile") == DETAIL_PROFILE # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : 예전 상세 규격 완료 체크포인트가 새 화면 수집을 건너뛰지 않도록 구분
    if not compatible:
        state = {}
    elif state.get("completed"):
        if args.mode == "full":
            return [], state, path
        state = {}
    last_id = state.get("lastId", "")
    if last_id:
        target_ids = [item["id"] for item in targets]
        if last_id in target_ids:
            targets = targets[target_ids.index(last_id) + 1:]
    state.setdefault("cycleId", datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ"))
    return targets, state, path


def save_checkpoint(path: Path | None, args: argparse.Namespace, state: dict[str, Any], processed: int, last_id: str, api_calls: int, completed: bool) -> None:
    if path is None:
        return
    write_json(path, {
        "schemaVersion": SCHEMA_VERSION, "updatedAt": now_iso(), "cycleId": state.get("cycleId"), "mode": args.mode,
        "detailProfile": DETAIL_PROFILE, # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : 상세 수집 규격 변경 시 샤드를 안전하게 새로 시작
        "scope": sorted(item.strip() for item in args.scope.split(",") if item.strip()),
        "shard": {"index": args.shard_index, "count": args.shard_count}, "processed": int(state.get("processed", 0)) + processed,
        "lastId": last_id, "apiCalls": int(state.get("apiCalls", 0)) + api_calls, "completed": completed,
    })


def main() -> int:
    load_local_env()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mode", choices=("incremental", "institution", "rotation", "full", "retry", "fixture"), default="incremental")
    parser.add_argument("--scope", default="all", help="catalog,details,photos,evaluations,all 중 쉼표 구분")
    parser.add_argument("--institution", action="append", default=[])
    parser.add_argument("--type", default="")
    parser.add_argument("--shard-count", type=int, default=14)
    parser.add_argument("--shard-index", type=int, default=0)
    parser.add_argument("--max-calls", type=int, default=MAX_DAILY_CALLS)
    parser.add_argument("--limit-institutions", type=int, default=0)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--refresh-source-files", action="store_true", help="공공데이터포털의 최신 시설현황·평가 원문을 먼저 확인")
    args = parser.parse_args()
    if args.shard_count < 1 or not 0 <= args.shard_index < args.shard_count:
        parser.error("shard index/count를 확인해 주세요.")
    scopes = {item.strip() for item in args.scope.split(",") if item.strip()}
    if "all" in scopes:
        scopes = {"catalog", "details", "photos", "evaluations"}
    invalid_scopes = scopes - {"catalog", "details", "photos", "evaluations"}
    if invalid_scopes:
        parser.error(f"지원하지 않는 scope: {', '.join(sorted(invalid_scopes))}")
    if args.mode == "fixture" and not args.institution:
        parser.error("fixture 모드는 --institution으로 검증 기관을 지정해야 합니다.") # SOFTM-NHIS-FIXTURE-SCOPE 날짜:20260902 : fixture가 샤드 전체에 빈 문서를 생성하지 않도록 범위를 제한

    if args.refresh_source_files:
        for source in FILE_SOURCES.values():
            path = refresh_source_file(source)
            print(f"공식 원문 확인: {path.name}")

    service_key = "" if args.mode == "fixture" else os.getenv("DATA_GO_KR_SERVICE_KEY", "").strip() # SOFTM-NHIS-FIXTURE 날짜:20260902 : 로컬 키가 있어도 fixture 모드는 네트워크를 호출하지 않음
    fixture_dir = FIXTURE_ROOT if args.mode == "fixture" or not service_key else None
    client = DataGoClient(service_key, min(MAX_DAILY_CALLS, args.max_calls), fixture_dir)
    old_catalog = load_json(DATA_ROOT / "catalog.json", {"institutions": []}).get("institutions", [])
    # /** SOFTM-NHIS-DETAIL-TARGET START 날짜:20260903 : 기관목록 수집 뒤 상세 수집이 같은 최종 기관목록 전체를 대상으로 이어지도록 보장 */
    catalog = build_xlsx_catalog() if "catalog" in scopes else old_catalog
    if not catalog:
        catalog = build_xlsx_catalog()
    # /** SOFTM-NHIS-DETAIL-TARGET END */
    live_seen: set[str] = set()
    if "catalog" in scopes and service_key and args.mode in {"incremental", "full"}:
        catalog, live_seen = merge_live_catalog(client, catalog)
    changes, changed_ids = catalog_changes(old_catalog, catalog, live_seen)
    generated_at = now_iso()
    for item in catalog:
        item["contentHash"] = catalog_hash(item)
    if "catalog" in scopes:
        write_json(DATA_ROOT / "catalog.json", {"schemaVersion": SCHEMA_VERSION, "generatedAt": generated_at, "source": ["data.go.kr:15059029", "data.go.kr:15124763"], "count": len(catalog), "institutions": catalog})
        write_json(DATA_ROOT / "changes.json", {"schemaVersion": SCHEMA_VERSION, "generatedAt": generated_at, **changes})
    evaluations_document = load_json(DATA_ROOT / "evaluations.json", {"institutions": {}})
    if "evaluations" in scopes:
        evaluations = evaluation_rows()
        evaluations_document = {"schemaVersion": SCHEMA_VERSION, "generatedAt": generated_at, "source": "data.go.kr:15104801", "count": len(evaluations), "institutions": evaluations}
        write_json(DATA_ROOT / "evaluations.json", evaluations_document)
    evaluations = evaluations_document.get("institutions", {})

    targets = select_institutions(args, catalog, changed_ids)
    targets, checkpoint_state, checkpoint_path = resume_shard(args, targets)
    failures: list[dict[str, str]] = []
    detail_success = photo_success = 0
    detail_updated = photo_updated = unchanged_count = processed_count = 0
    last_processed_id = checkpoint_state.get("lastId", "")
    budget_exhausted = False
    for index, institution in enumerate(targets, 1):
        service_code = preferred_service(institution, args.type.upper())
        if not service_code:
            failures.append({"id": institution["id"], "scope": "details", "message": "급여종류 코드 없음"})
            processed_count += 1
            last_processed_id = institution["id"]
            continue
        shard = institution["id"][:2]
        if "details" in scopes:
            path = DATA_ROOT / "details" / shard / f"{institution['id']}.json.gz"
            legacy_path = path.with_suffix("")
            existing = load_json(path, {}) if path.exists() else load_json(legacy_path, {})  # SOFTM-NHIS-GZIP 날짜:20260903 : 압축 전환 중에도 기존 JSON 체크포인트 결과를 이어받기 위한 호환 읽기
            requested_codes = [args.type.upper()] if args.type and args.type.upper() in [service["code"] for service in institution.get("services", [])] else [service["code"] for service in institution.get("services", [])]
            # /** SOFTM-NHIS-OFFICIAL-PAGE START 날짜:20260903 : 화면 탭 보완 전의 기존 상세 JSON은 완료로 오인하지 않고 순차 갱신 */
            complete = bool(requested_codes) and existing.get("id") == institution["id"] and all(
                code in existing.get("serviceDetails", {})
                and not existing["serviceDetails"][code].get("failures")
                and set(DETAIL_OPERATIONS).issubset(existing["serviceDetails"][code].get("sections", {}))
                and (args.mode == "fixture" or set(map(str, OFFICIAL_DETAIL_TABS)).issubset(existing["serviceDetails"][code].get("officialPage", {}).get("tabs", {})))
                for code in requested_codes
            )
            # /** SOFTM-NHIS-OFFICIAL-PAGE END */
            if complete and not args.force and args.mode not in {"rotation", "full"} and institution["id"] not in changed_ids:
                detail_success += 1
                unchanged_count += 1
            else:
                try:
                    document = existing
                    for code in requested_codes or [service_code]:
                        detail = fetch_detail(client, institution, code, include_official_page=args.mode != "fixture") # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : fixture는 외부 화면을 호출하지 않고 실제 수집만 보완
                        document = merge_detail_document(document, institution, detail, evaluations)
                        for item in detail.get("failures", []):
                            failures.append({"id": institution["id"], "scope": "details", "serviceCode": code, "section": item.get("section", ""), "message": item.get("message", "")})
                    write_gzip_json(path, document)  # SOFTM-NHIS-GZIP 날짜:20260903 : 상세 생성 단계부터 압축해 비압축 파일이 다시 누적되지 않도록 저장
                    detail_success += 1
                    detail_updated += 1
                except ApiBudgetExceeded as error:
                    failures.append({"id": institution["id"], "scope": "details", "message": str(error)})
                    budget_exhausted = True
                    break
                except Exception as error:
                    failures.append({"id": institution["id"], "scope": "details", "message": str(error)[:300]})
        if "photos" in scopes:
            path = DATA_ROOT / "photos" / shard / f"{institution['id']}.json"
            if path.exists() and not args.force and args.mode not in {"rotation", "full"} and institution["id"] not in changed_ids:
                photo_success += 1
                unchanged_count += 1
            else:
                try:
                    max_photos = min(30, max(1, int(os.getenv("NHIS_MAX_PHOTOS_PER_INSTITUTION", "10"))))
                    photo_mode = os.getenv("NHIS_PHOTO_MODE", "remote").strip().lower()
                    write_json(path, collect_photos(institution, service_code, max_photos, photo_mode))
                    photo_success += 1
                    photo_updated += 1
                    time.sleep(random.uniform(0.15, 0.45))
                except Exception as error:
                    failures.append({"id": institution["id"], "scope": "photos", "message": str(error)[:300]})
        processed_count += 1
        last_processed_id = institution["id"]
        if index % 25 == 0:
            save_checkpoint(checkpoint_path, args, checkpoint_state, processed_count, last_processed_id, client.calls, False)

    shard_completed = not budget_exhausted and processed_count == len(targets)
    save_checkpoint(checkpoint_path, args, checkpoint_state, processed_count, last_processed_id, client.calls, shard_completed)

    detail_failures = [item for item in failures if item["scope"] == "details"]
    photo_failures = [item for item in failures if item["scope"] == "photos"]
    if "details" in scopes:
        write_json(DATA_ROOT / "failures" / "details.json", {"schemaVersion": SCHEMA_VERSION, "generatedAt": generated_at, "items": detail_failures})
    if "photos" in scopes:
        write_json(DATA_ROOT / "failures" / "photos.json", {"schemaVersion": SCHEMA_VERSION, "generatedAt": generated_at, "items": photo_failures})
    manifest = load_json(DATA_ROOT / "manifest.json", {})
    completed_shards = []
    for checkpoint_file in ((DATA_ROOT / "checkpoints").glob(f"{args.mode}-*.json") if (DATA_ROOT / "checkpoints").exists() else []):
        checkpoint = load_json(checkpoint_file, {})
        if checkpoint.get("completed") and checkpoint.get("shard", {}).get("count") == args.shard_count:
            completed_shards.append(checkpoint["shard"]["index"])
    detail_ids = sorted(path.name.removesuffix(".json.gz") for path in (DATA_ROOT / "details").glob("*/*.json.gz")) if (DATA_ROOT / "details").exists() else []  # SOFTM-NHIS-GZIP 날짜:20260903 : 매니페스트가 실제 압축 상세 파일을 기준으로 기관기호를 기록
    photo_ids = sorted(path.stem for path in (DATA_ROOT / "photos").glob("*/*.json")) if (DATA_ROOT / "photos").exists() else []
    manifest.update({
        "schemaVersion": SCHEMA_VERSION, "detailProfile": DETAIL_PROFILE, "cycleId": checkpoint_state.get("cycleId") or datetime.now(timezone.utc).strftime("%Y%m"), "mode": args.mode, "generatedAt": generated_at, "updatedAt": generated_at, "catalogCount": len(catalog), # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : 배포 데이터가 공단 화면 보완 규격인지 확인 가능하게 기록
        "detailCount": len(detail_ids), "detailIds": detail_ids,
        "photoManifestCount": len(photo_ids), "photoIds": photo_ids,
        "evaluationCount": load_json(DATA_ROOT / "evaluations.json", {}).get("count", 0),
        "completedShards": sorted(set(completed_shards)),
        "lastRun": {"mode": args.mode, "scope": sorted(scopes), "targets": len(targets), "processed": processed_count, "detailSuccess": detail_success, "photoSuccess": photo_success, "updated": detail_updated + photo_updated, "unchanged": unchanged_count, "failures": len(failures), "apiCalls": client.calls, "shardIndex": args.shard_index, "shardCount": args.shard_count},
        "updatedCount": detail_updated + photo_updated, "unchangedCount": unchanged_count, "failureCount": len(failures),
        "sources": {
            "catalogApi": "https://www.data.go.kr/data/15059029/openapi.do", "detailApi": "https://www.data.go.kr/data/15058856/openapi.do",
            "facilityFile": "https://www.data.go.kr/data/15124763/fileData.do", "evaluationFile": "https://www.data.go.kr/data/15104801/fileData.do",
            "officialDetailPage": PHOTO_PAGE, # SOFTM-NHIS-OFFICIAL-PAGE 날짜:20260903 : 기본·근속·CCTV 화면 항목의 실제 수집 출처
        },
    })
    write_json(DATA_ROOT / "manifest.json", manifest)
    print(json.dumps(manifest["lastRun"], ensure_ascii=False, indent=2))
    return 1 if failures and not (detail_success or photo_success) else 0


if __name__ == "__main__":
    sys.exit(main())
# /** SOFTM-NHIS-STATIC END */
